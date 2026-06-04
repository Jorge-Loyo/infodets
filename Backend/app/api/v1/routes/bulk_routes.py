"""
HU-029: Importación masiva de usuarios desde CSV
HU-030: Exportar reportes de uso en Excel
"""
import io
import csv
import logging
from datetime import datetime, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, cast, Date

from app.core.database import get_db
from app.core.settings import settings
from app.middleware.auth_middleware import require_permiso
from app.models.models import (
    HistorialChat, TicketVacio, ConsultaInvitado,
    Documento, Usuario, ReporteFeedback,
)
from app.services import usuario_service, perfil_service
from app.services import audit_service
from app.services.notificacion_service import notificar_admin_sync

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/admin/bulk", tags=["Importación y Exportación"])


# ── HU-029: Importación masiva CSV ────────────────────────────────────────────

class ImportResult:
    def __init__(self):
        self.exitosos: list[str] = []
        self.errores: list[dict] = []


@router.post(
    "/importar-usuarios",
    status_code=200,
    summary="Importar usuarios desde archivo CSV",
    description=(
        "Acepta un CSV con columnas: email, nombre, apellido, perfil_id (obligatorias). "
        "Opcionales: dni, cargo, institucion, dependencia. "
        "Retorna reporte de éxitos y errores por fila."
    ),
)
async def importar_usuarios_csv(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_permiso("gestionar_usuarios")),
):
    if not archivo.filename or not archivo.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos CSV")

    contenido = await archivo.read()
    try:
        texto = contenido.decode("utf-8-sig")
    except UnicodeDecodeError:
        texto = contenido.decode("latin-1")

    reader = csv.DictReader(io.StringIO(texto))
    resultado = ImportResult()

    campos_requeridos = {"email", "nombre", "apellido", "perfil_id"}
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="El archivo CSV está vacío o no tiene encabezados")

    campos_csv = {f.strip().lower() for f in reader.fieldnames}
    faltantes = campos_requeridos - campos_csv
    if faltantes:
        raise HTTPException(
            status_code=400,
            detail=f"Faltan columnas obligatorias: {', '.join(faltantes)}. Columnas encontradas: {', '.join(reader.fieldnames)}"
        )

    for fila_num, row in enumerate(reader, start=2):
        # Normalizar claves
        row = {k.strip().lower(): (v.strip() if v else "") for k, v in row.items()}
        email = row.get("email", "").lower()
        nombre = row.get("nombre", "")
        apellido = row.get("apellido", "")
        perfil_id = row.get("perfil_id", "")

        # Validaciones
        if not email or "@" not in email:
            resultado.errores.append({"fila": fila_num, "email": email, "error": "Email inválido o vacío"})
            continue
        if not nombre:
            resultado.errores.append({"fila": fila_num, "email": email, "error": "Nombre vacío"})
            continue
        if not perfil_id:
            resultado.errores.append({"fila": fila_num, "email": email, "error": "perfil_id vacío"})
            continue

        # Verificar duplicados
        existente = usuario_service.obtener_usuario_por_email(db, email)
        if existente:
            resultado.errores.append({"fila": fila_num, "email": email, "error": "Usuario ya existe"})
            continue

        # Verificar perfil válido
        perfil = perfil_service.obtener_perfil(db, perfil_id)
        if not perfil:
            resultado.errores.append({"fila": fila_num, "email": email, "error": f"perfil_id '{perfil_id}' no existe"})
            continue

        try:
            usuario = usuario_service.invitar_usuario(
                db,
                email=email,
                nombre=nombre,
                apellido=apellido,
                rol="operador",
                dni=row.get("dni"),
                fecha_nacimiento=row.get("fecha_nacimiento"),
                cargo=row.get("cargo"),
                institucion=row.get("institucion"),
                dependencia=row.get("dependencia"),
                perfil_id=perfil_id,
            )
            perfil_service.asignar_perfil_a_usuario(db, str(usuario.id), perfil_id)
            resultado.exitosos.append(email)
        except Exception as e:
            resultado.errores.append({"fila": fila_num, "email": email, "error": str(e)})

    # Auditoría
    audit_service.registrar(
        db, accion="importar_csv", entidad="usuario",
        entidad_id=None,
        detalle=f"Importación CSV: {len(resultado.exitosos)} exitosos, {len(resultado.errores)} errores",
        realizado_por_id=current_user.get("_usuario_id"),
        realizado_por_email=current_user.get("email"),
    )

    return {
        "total_procesados": len(resultado.exitosos) + len(resultado.errores),
        "exitosos": len(resultado.exitosos),
        "errores_count": len(resultado.errores),
        "usuarios_creados": resultado.exitosos,
        "errores": resultado.errores,
    }


@router.get(
    "/template-csv",
    summary="Descargar plantilla CSV para importación",
    description="Retorna un CSV de ejemplo con los encabezados correctos.",
)
async def descargar_template_csv(
    current_user: dict = Depends(require_permiso("gestionar_usuarios")),
):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["email", "nombre", "apellido", "perfil_id", "dni", "cargo", "institucion", "dependencia"])
    writer.writerow(["ejemplo@entidad.gob.ar", "Juan", "Pérez", "uuid-del-perfil", "12345678", "Analista", "ANSES", "Sistemas"])
    output.seek(0)

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=plantilla_usuarios.csv"},
    )


# ── HU-030: Exportar reportes Excel ──────────────────────────────────────────

@router.get(
    "/exportar/consultas",
    summary="Exportar historial de consultas en Excel",
    description="Exporta consultas filtradas por fecha. Formato XLSX.",
)
async def exportar_consultas(
    desde: Optional[str] = Query(None, description="Fecha inicio ISO (ej: 2025-06-01)"),
    hasta: Optional[str] = Query(None, description="Fecha fin ISO (ej: 2025-06-30)"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_permiso("dashboard")),
):
    from openpyxl import Workbook

    query = db.query(HistorialChat).order_by(HistorialChat.creado_en.desc())

    if desde:
        query = query.filter(HistorialChat.creado_en >= datetime.fromisoformat(desde))
    if hasta:
        query = query.filter(HistorialChat.creado_en <= datetime.fromisoformat(hasta) + timedelta(days=1))

    registros = query.limit(5000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Consultas"
    ws.append(["ID", "Usuario ID", "Pregunta", "Respuesta", "Confianza", "Es Fallback", "Fecha"])

    for r in registros:
        ws.append([
            str(r.id),
            str(r.usuario_id),
            r.pregunta[:500],
            r.respuesta[:500],
            round(r.puntaje_confianza, 3),
            "Sí" if r.es_fallback else "No",
            r.creado_en.strftime("%Y-%m-%d %H:%M") if r.creado_en else "",
        ])

    # Autoajustar anchos
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 60)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=consultas_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )


@router.get(
    "/exportar/tickets",
    summary="Exportar tickets en Excel",
    description="Exporta tickets filtrados por estado y fecha.",
)
async def exportar_tickets(
    estado: Optional[str] = Query(None, description="pendiente | revisado | respondido"),
    desde: Optional[str] = Query(None),
    hasta: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_permiso("dashboard")),
):
    from openpyxl import Workbook

    query = db.query(TicketVacio).order_by(TicketVacio.creado_en.desc())

    if estado:
        query = query.filter(TicketVacio.estado == estado)
    if desde:
        query = query.filter(TicketVacio.creado_en >= datetime.fromisoformat(desde))
    if hasta:
        query = query.filter(TicketVacio.creado_en <= datetime.fromisoformat(hasta) + timedelta(days=1))

    tickets = query.limit(5000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    ws.append(["ID", "Pregunta", "Usuario", "Confianza", "Nivel", "Estado", "Requiere Respuesta", "Fecha"])

    for t in tickets:
        ws.append([
            str(t.id),
            t.pregunta[:500],
            t.usuario_id or "",
            round(t.puntaje_confianza, 3),
            t.nivel,
            t.estado,
            "Sí" if t.requiere_respuesta else "No",
            t.creado_en.strftime("%Y-%m-%d %H:%M") if t.creado_en else "",
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 60)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=tickets_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )


@router.get(
    "/exportar/usuarios",
    summary="Exportar usuarios en Excel",
)
async def exportar_usuarios(
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_permiso("gestionar_usuarios")),
):
    from openpyxl import Workbook

    usuarios = db.query(Usuario).order_by(Usuario.creado_en.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    ws.append(["ID", "Email", "Nombre", "Apellido", "DNI", "Cargo", "Institución", "Dependencia", "Rol", "Creado"])

    for u in usuarios:
        ws.append([
            str(u.id),
            u.email,
            u.nombre or "",
            u.apellido or "",
            u.dni or "",
            u.cargo or "",
            u.institucion or "",
            u.dependencia or "",
            u.rol.value if hasattr(u.rol, 'value') else u.rol,
            u.creado_en.strftime("%Y-%m-%d %H:%M") if u.creado_en else "",
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=usuarios_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )
